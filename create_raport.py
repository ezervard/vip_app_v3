import openpyxl
import datetime
class CreateRaport:
    def __init__(self):#Инициализация класовых переменных
        self.file_path = 'BRAMY_VIP_MG2.xlsx'
        self.workbook = None
        self.worksheet = None
        self.current_datetime = datetime.datetime.now()
        self.act = self.current_datetime.strftime('%d.%m.%Y')



    def raport(self):# Инициализация файла эксель
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.worksheet = self.workbook.active
        except FileNotFoundError:
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
        except Exception as e:
            print(f"Произошла ошибка при открытии файла: {e}")

    def get_raport(self, raport, entr, row_index):#Функция заполнения рапорта
        if self.workbook is None:
            self.raport()
        #Временное хранение уже записанных данных
        plate_db = []
        time_entr = []
        time_exit = []

        #Перебор строк в рапорте и добавление уже существующих данных во временные переменные
        for row in self.worksheet.iter_rows():
            plates_data = row[6].value
            exit_data = row[9].value
            entr_data = row[7].value
            if row is not None:
                time_exit.append(exit_data)
                plate_db.append(plates_data)
                time_entr.append(entr_data)

        print(plate_db)
        print(time_entr)

        #Если документ уже сожержит значения введеные пользователем то просто добавляется время
        if raport[3] in plate_db:
            index = plate_db.index(raport[3])

            if entr == True: #если отработала функция entrance время добавляется в колонку вьезд
                if time_entr[index] :
                    self.worksheet[f'H{index +1}'] = f'{time_entr[index]}/{raport[5]}'
                    self.worksheet[f'J{index + 1}'] = f'{time_exit[index]}'


            elif entr == False: #если отработала функция exit время добавляется в колонку выезд
                if time_exit[index] :
                    self.worksheet[f'J{index + 1}'] = f'{time_exit[index]}/{raport[5]}'
                    self.worksheet[f'H{index + 1}'] = f'{time_entr[index]}'

        else: #Если номера еще нет то добавляются все значения

            data = raport  # Присваиваем data значение из списка из базы данных
            self.worksheet[f'D{row_index}'] = data[0]  # Тип пропуска
            self.worksheet[f'E{row_index}'] = data[1]  # Тип машины
            self.worksheet[f'F{row_index}'] = data[2] #Фирма
            self.worksheet[f'G{row_index}'] = data[3] #Номер машины
            if entr:#Проверяем если машина вьезжает то время доавляется в колонку вьезд если нет то в выезд
                self.worksheet[f'H{row_index}'] = data[5]
            else:
                self.worksheet[f'J{row_index}'] = data[5]
            self.worksheet[f'I{row_index}'] = data[4] #Добавляем место паркавки
            self.worksheet[f'K{row_index}'] = data[6] # Добавляем контроль багажника
            self.worksheet[f'L{row_index}'] = data[7] # Добавляем Имя и Фамилию
            return row_index + 1





        try:
            self.workbook.save(self.act + '_BRAMY_VIP_MG2.xlsx')  # Сохраняем изменения в файл
        except PermissionError:
            print("файл открыт в другой программе")






    def save_workbook(self):
        self.workbook.save(self.act +'_BRAMY_VIP_MG2.xlsx')
        print('saved')
        print(self.current_datetime.strftime('%d.%m.%Y'))
        self.workbook.close()